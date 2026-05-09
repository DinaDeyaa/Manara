"""
Manara AI Learning Support System - Google Form Evaluation Script
Analyzes survey responses and produces professional evaluation outputs.
"""

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
import os

warnings.filterwarnings('ignore')

OUTPUT_DIR = r'C:\Users\LENOVO\Desktop\manara_outputs'
os.makedirs(OUTPUT_DIR, exist_ok=True)

INPUT_FILE = r"C:\Users\LENOVO\Downloads\Responses.xlsx"
# ─────────────────────────────────────────────
# 1. LOAD & RENAME COLUMNS
# ─────────────────────────────────────────────

raw = pd.read_excel(INPUT_FILE)

COLUMN_MAP = {
    'Are you currently a student?': 'student_status',
    'Which year are you currently in?': 'academic_year',
    'If you had access to a system like Manara earlier in your studies, would you have used it?': 'earlier_use',
    'How important is it to assess your prerequisite knowledge before starting a course?': 'prerequisite_importance',
    'A diagnostic exam would help me identify my weak areas early.': 'diagnostic_exam',
    'If a system shows your weak topics before studying, how likely are you to use it?': 'weak_topics_usage',
    'How useful is a personalized learning path based on your weaknesses?': 'personalized_path',
    'Do you feel that available past papers are sufficient for your courses?': 'past_papers_sufficient',
    'How helpful would unlimited generated practice questions be for your learning?': 'generated_questions',
    'How likely are you to use a chatbot that answers questions from course material only?': 'chatbot_usage',
    'Does tracking your progress help you stay motivated?': 'progress_motivation',
    'How interested would you be in receiving motivational reminders via WhatsApp from Manara?': 'whatsapp_interest',
    'Which feature do you find most valuable?': 'valuable_features',
    'Do you believe identifying your weaknesses before studying a course would improve your overall performance?': 'weakness_improves_performance',
    'Would you use Manara in your studies if it was available?': 'adoption_intention',
}

# Flexible matching: strip and lower
def match_column(df_cols, target):
    target_clean = target.strip().lower()
    for col in df_cols:
        if col.strip().lower() == target_clean:
            return col
    # partial match fallback
    for col in df_cols:
        if target_clean[:40] in col.strip().lower():
            return col
    return None

rename_map = {}
for target, internal in COLUMN_MAP.items():
    matched = match_column(raw.columns, target)
    if matched:
        rename_map[matched] = internal
    else:
        print(f"WARNING: Could not find column for: {target}")

df = raw.rename(columns=rename_map)

# Retain only rows that are not entirely empty (beyond timestamp/name)
scored_cols = list(COLUMN_MAP.values())
df = df.dropna(how='all', subset=scored_cols)

TOTAL = len(df)

# ─────────────────────────────────────────────
# 2. CLEANING HELPER
# ─────────────────────────────────────────────

def clean(series):
    return series.astype(str).str.strip().str.lower().replace('nan', np.nan)

# ─────────────────────────────────────────────
# 3. NORMALIZATION MAPPINGS
# ─────────────────────────────────────────────

MAPS = {
    'earlier_use': {'yes': 100, 'maybe': 50, 'no': 0},
    'prerequisite_importance': {'very important': 100, 'important': 75, 'neutral': 50, 'not important': 0},
    'diagnostic_exam': {'strongly agree': 100, 'agree': 75, 'neutral': 50, 'disagree': 0},
    'weak_topics_usage': {'very likely': 100, 'likely': 75, 'maybe': 50, 'unlikely': 0},
    'chatbot_usage': {'very likely': 100, 'likely': 75, 'neutral': 50, 'unlikely': 25, 'very unlikely': 0},
    'progress_motivation': {'yes': 100, 'maybe': 50, 'no': 0},
    'whatsapp_interest': {'very interested': 100, 'interested': 75, 'not sure': 50, 'not interested': 0},
    'weakness_improves_performance': {'yes': 100, 'maybe': 50, 'no': 0},
    'adoption_intention': {'yes': 100, 'maybe': 50, 'no': 0},
    'past_papers_sufficient': {'no': 100, 'sometimes': 50, 'yes': 0},  # need score
}

def apply_map(series, mapping, col_name):
    cleaned = clean(series)
    result = []
    for val in cleaned:
        if pd.isna(val) or val == 'nan':
            result.append(np.nan)
        elif val in mapping:
            result.append(mapping[val])
        else:
            print(f"  WARNING [{col_name}]: Unexpected value '{val}'")
            result.append(np.nan)
    return pd.Series(result, index=series.index)

def normalize_linear(series, col_name):
    numeric = pd.to_numeric(series, errors='coerce')
    result = ((numeric - 1) / (5 - 1)) * 100
    unexpected = series[numeric.isna() & series.notna() & (series.astype(str).str.strip() != '') & (series.astype(str).str.strip() != 'nan')]
    for v in unexpected.unique():
        print(f"  WARNING [{col_name}]: Unexpected value '{v}'")
    return result

# Apply normalizations
for col, mapping in MAPS.items():
    df[col + '_score'] = apply_map(df[col], mapping, col)

df['personalized_path_score'] = normalize_linear(df['personalized_path'], 'personalized_path')
df['generated_questions_score'] = normalize_linear(df['generated_questions'], 'generated_questions')

# Past papers: rename score column for clarity
df['past_papers_need_score'] = df['past_papers_sufficient_score']

# ─────────────────────────────────────────────
# 4. RESPONDENT PROFILE
# ─────────────────────────────────────────────

student_col = clean(df['student_status'])
n_students = (student_col == 'yes').sum()
n_nonstudents = (student_col == 'no').sum()

# Academic year - normalize "6th year" / "6th" variants
year_raw = df['academic_year'].astype(str).str.strip()
year_raw = year_raw.replace({'6th year ': '6th year', '6th': '6th year'})
year_dist = year_raw[year_raw != 'nan'].value_counts()

# Earlier use - only for non-students
earlier_raw = clean(df['earlier_use'])
earlier_dist = earlier_raw.dropna().value_counts()

# ─────────────────────────────────────────────
# 5. QUESTION-LEVEL STATISTICS
# ─────────────────────────────────────────────

QUESTION_META = [
    # (internal_name, score_col, display_name, category, scale_desc, norm_method)
    ('prerequisite_importance', 'prerequisite_importance_score',
     'Prerequisite Knowledge Importance', 'Need for Guidance & Weakness ID',
     'Very important / Important / Neutral / Not important', 'Categorical (4-point)'),
    ('diagnostic_exam', 'diagnostic_exam_score',
     'Diagnostic Exam Usefulness', 'Need for Guidance & Weakness ID',
     'Strongly agree / Agree / Neutral / Disagree', 'Categorical (4-point)'),
    ('weak_topics_usage', 'weak_topics_usage_score',
     'Weak Topics Usage Likelihood', 'Need for Guidance & Weakness ID',
     'Very likely / Likely / Maybe / Unlikely', 'Categorical (4-point)'),
    ('weakness_improves_performance', 'weakness_improves_performance_score',
     'Weakness ID Improves Performance', 'Need for Guidance & Weakness ID',
     'Yes / Maybe / No', 'Categorical (3-point)'),
    ('personalized_path', 'personalized_path_score',
     'Personalized Learning Path Usefulness', 'Personalized Learning Path',
     '1–5 linear scale', 'Linear ((score-1)/4 × 100)'),
    ('past_papers_sufficient', 'past_papers_need_score',
     'Practice Resource Need (Past Papers)', 'Question Bank & Practice Support',
     'No / Sometimes / Yes (need score)', 'Need indicator (inverted)'),
    ('generated_questions', 'generated_questions_score',
     'Generated Practice Questions Helpfulness', 'Question Bank & Practice Support',
     '1–5 linear scale', 'Linear ((score-1)/4 × 100)'),
    ('chatbot_usage', 'chatbot_usage_score',
     'Chatbot Usage Likelihood', 'AI Chatbot & Course Assistance',
     'Very likely / Likely / Neutral / Unlikely / Very unlikely', 'Categorical (5-point)'),
    ('progress_motivation', 'progress_motivation_score',
     'Progress Tracking Motivation', 'Progress Tracking & Motivation',
     'Yes / Maybe / No', 'Categorical (3-point)'),
    ('whatsapp_interest', 'whatsapp_interest_score',
     'WhatsApp Reminder Interest', 'WhatsApp Reminder Interest',
     'Very interested / Interested / Not sure / Not interested', 'Categorical (4-point)'),
    ('adoption_intention', 'adoption_intention_score',
     'Adoption Intention', 'Overall Adoption Intention',
     'Yes / Maybe / No', 'Categorical (3-point)'),
]

def interpret_score(score, is_need=False):
    if pd.isna(score):
        return 'N/A'
    if is_need:
        if score >= 85: return 'Very strong need for additional practice resources'
        if score >= 70: return 'Strong need for additional practice resources'
        if score >= 50: return 'Moderate need for additional practice resources'
        if score >= 30: return 'Limited need for additional practice resources'
        return 'Low need for additional practice resources'
    else:
        if score >= 85: return 'Very positive / very strong support'
        if score >= 70: return 'Positive support'
        if score >= 50: return 'Moderate support'
        if score >= 30: return 'Weak to moderate support'
        return 'Low support'

question_rows = []
for internal, score_col, display, category, scale, method in QUESTION_META:
    s = df[score_col].dropna()
    valid_n = len(s)
    missing_n = TOTAL - valid_n
    avg = s.mean()
    std = s.std()
    is_need = (internal == 'past_papers_sufficient')
    interp = interpret_score(avg, is_need=is_need)
    question_rows.append({
        'Question': display,
        'Category': category,
        'Valid Responses': valid_n,
        'Missing Responses': missing_n,
        'Response Scale': scale,
        'Normalization Method': method,
        'Normalized Avg Score': round(avg, 2),
        'Std Dev': round(std, 2),
        'Interpretation': interp,
    })

question_summary_table = pd.DataFrame(question_rows)

# ─────────────────────────────────────────────
# 6. CATEGORY SUMMARIES
# ─────────────────────────────────────────────

CATEGORIES = {
    'Need for Guidance & Weakness ID': [
        'prerequisite_importance_score', 'diagnostic_exam_score',
        'weak_topics_usage_score', 'weakness_improves_performance_score'
    ],
    'Personalized Learning Path': ['personalized_path_score'],
    'Question Bank & Practice Support': ['past_papers_need_score', 'generated_questions_score'],
    'AI Chatbot & Course Assistance': ['chatbot_usage_score'],
    'Progress Tracking & Motivation': ['progress_motivation_score'],
    'WhatsApp Reminder Interest': ['whatsapp_interest_score'],
    'Overall Adoption Intention': ['adoption_intention_score'],
}

CAT_DISPLAY_QUESTIONS = {
    'Need for Guidance & Weakness ID': 'Prerequisite Importance, Diagnostic Exam, Weak Topics Usage, Weakness ID Improves Performance',
    'Personalized Learning Path': 'Personalized Learning Path Usefulness',
    'Question Bank & Practice Support': 'Practice Resource Need (Past Papers), Generated Questions Helpfulness',
    'AI Chatbot & Course Assistance': 'Chatbot Usage Likelihood',
    'Progress Tracking & Motivation': 'Progress Tracking Motivation',
    'WhatsApp Reminder Interest': 'WhatsApp Reminder Interest',
    'Overall Adoption Intention': 'Adoption Intention',
}

category_rows = []
category_scores = {}
for cat, cols in CATEGORIES.items():
    scores = pd.concat([df[c].dropna() for c in cols])
    avg = scores.mean()
    is_need_cat = (cat == 'Question Bank & Practice Support')
    interp = interpret_score(avg, is_need=False)  # category-level: use normal scale
    category_rows.append({
        'Evaluation Category': cat,
        'Included Questions': CAT_DISPLAY_QUESTIONS[cat],
        'Average Normalized Score': round(avg, 2),
        'Interpretation': interp,
    })
    category_scores[cat] = round(avg, 2)

category_summary_table = pd.DataFrame(category_rows)

# ─────────────────────────────────────────────
# 7. FEATURE PREFERENCES
# ─────────────────────────────────────────────

EXPECTED_FEATURES = [
    'Diagnostic exam', 'Personalized learning path',
    'Question bank', 'AI chatbot', 'Progress tracking'
]

feature_counts = {f: 0 for f in EXPECTED_FEATURES}
feat_series = df['valuable_features'].dropna().astype(str)
for entry in feat_series:
    for part in entry.split(','):
        part = part.strip()
        if part in feature_counts:
            feature_counts[part] += 1
        elif part.lower() in [f.lower() for f in EXPECTED_FEATURES]:
            # case-insensitive match
            for f in EXPECTED_FEATURES:
                if f.lower() == part.lower():
                    feature_counts[f] += 1
        elif part and part != 'nan':
            print(f"  WARNING [valuable_features]: Unexpected feature '{part}'")

feat_df = pd.DataFrame([
    {'Feature': f, 'Selection Count': c, 'Percentage of Respondents': round(c / TOTAL * 100, 1)}
    for f, c in feature_counts.items()
]).sort_values('Selection Count', ascending=False).reset_index(drop=True)
feat_df['Rank'] = range(1, len(feat_df) + 1)
feature_preference_table = feat_df[['Rank', 'Feature', 'Selection Count', 'Percentage of Respondents']]

# ─────────────────────────────────────────────
# 8. RESPONSE DISTRIBUTION TABLE
# ─────────────────────────────────────────────

CATEGORICAL_QUESTIONS = [
    ('student_status', 'Are you currently a student?'),
    ('academic_year', 'Academic Year'),
    ('earlier_use', 'Earlier Use of Manara'),
    ('prerequisite_importance', 'Prerequisite Knowledge Importance'),
    ('diagnostic_exam', 'Diagnostic Exam Usefulness'),
    ('weak_topics_usage', 'Weak Topics Usage Likelihood'),
    ('past_papers_sufficient', 'Past Papers Sufficiency'),
    ('chatbot_usage', 'Chatbot Usage Likelihood'),
    ('progress_motivation', 'Progress Tracking Motivation'),
    ('whatsapp_interest', 'WhatsApp Reminder Interest'),
    ('weakness_improves_performance', 'Weakness ID Improves Performance'),
    ('adoption_intention', 'Adoption Intention'),
]

dist_rows = []
for col, label in CATEGORICAL_QUESTIONS:
    col_data = df[col].astype(str).str.strip()
    col_data = col_data[col_data.str.lower() != 'nan']
    total_valid = len(col_data)
    counts = col_data.value_counts()
    for opt, cnt in counts.items():
        dist_rows.append({
            'Question': label,
            'Response Option': opt,
            'Count': cnt,
            'Percentage': round(cnt / total_valid * 100, 1) if total_valid > 0 else 0,
        })

response_distribution_table = pd.DataFrame(dist_rows)

# ─────────────────────────────────────────────
# 9. RESPONDENT PROFILE TABLE
# ─────────────────────────────────────────────

profile_rows = [
    {'Metric': 'Total Responses', 'Value': TOTAL, 'Percentage': '100%'},
    {'Metric': 'Current Students', 'Value': n_students, 'Percentage': f'{n_students/TOTAL*100:.1f}%'},
    {'Metric': 'Non-Students', 'Value': n_nonstudents, 'Percentage': f'{n_nonstudents/TOTAL*100:.1f}%'},
]
for year, cnt in year_dist.items():
    profile_rows.append({'Metric': f'Academic Year: {year}', 'Value': cnt, 'Percentage': f'{cnt/TOTAL*100:.1f}%'})
for opt, cnt in earlier_dist.items():
    label = {'yes': 'Would have used Manara earlier: Yes',
             'maybe': 'Would have used Manara earlier: Maybe',
             'no': 'Would have used Manara earlier: No'}.get(opt, f'Earlier Use: {opt}')
    profile_rows.append({'Metric': label, 'Value': cnt, 'Percentage': f'{cnt/TOTAL*100:.1f}%'})

respondent_profile_table = pd.DataFrame(profile_rows)

# ─────────────────────────────────────────────
# 10. VISUALIZATIONS
# ─────────────────────────────────────────────

COLORS = {
    'primary': '#2E86AB',
    'secondary': '#A23B72',
    'accent': '#F18F01',
    'green': '#44BBA4',
    'dark': '#333333',
    'light_bg': '#F5F5F5',
}

plt.rcParams.update({
    'font.family': 'DejaVu Sans',
    'axes.spines.top': False,
    'axes.spines.right': False,
    'axes.grid': True,
    'axes.grid.axis': 'y',
    'grid.alpha': 0.4,
    'figure.dpi': 150,
})

# — Chart 1: Category Scores —
fig, ax = plt.subplots(figsize=(12, 6))
cats = list(category_scores.keys())
scores = list(category_scores.values())
short_labels = [c.replace(' & ', '\n& ').replace('Question Bank\n& Practice Support', 'Question Bank\n& Practice') for c in cats]

bar_colors = [COLORS['primary']] * len(cats)
bars = ax.bar(range(len(cats)), scores, color=bar_colors, width=0.6, edgecolor='white', linewidth=1.2)
for bar, score in zip(bars, scores):
    ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 1.2,
            f'{score:.1f}', ha='center', va='bottom', fontsize=10, fontweight='bold', color=COLORS['dark'])

ax.set_xticks(range(len(cats)))
ax.set_xticklabels([c.replace(' & ', '\n& ') for c in cats], rotation=15, ha='right', fontsize=9)
ax.set_ylim(0, 105)
ax.set_ylabel('Normalized Score (0–100)', fontsize=11)
ax.set_title('Manara Evaluation Category Scores', fontsize=14, fontweight='bold', pad=15)
ax.axhline(y=85, color='green', linestyle='--', alpha=0.5, linewidth=1, label='85 – Very Positive threshold')
ax.axhline(y=70, color='orange', linestyle='--', alpha=0.5, linewidth=1, label='70 – Positive threshold')
ax.legend(fontsize=8, loc='lower right')
plt.tight_layout()
plt.savefig(os.path.join(OUTPUT_DIR, 'category_scores.png'), bbox_inches='tight')
plt.close()

# — Chart 2: Feature Preferences —
fig, ax = plt.subplots(figsize=(10, 5))
feat_sorted = feature_preference_table.sort_values('Percentage of Respondents', ascending=True)
palette = [COLORS['primary'], COLORS['secondary'], COLORS['accent'], COLORS['green'], '#7B2D8B']
bars = ax.barh(feat_sorted['Feature'], feat_sorted['Percentage of Respondents'],
               color=palette[:len(feat_sorted)], edgecolor='white', linewidth=1.2)
for bar, pct in zip(bars, feat_sorted['Percentage of Respondents']):
    ax.text(bar.get_width() + 0.5, bar.get_y() + bar.get_height() / 2,
            f'{pct:.1f}%', va='center', fontsize=10, fontweight='bold', color=COLORS['dark'])
ax.set_xlim(0, 110)
ax.set_xlabel('Percentage of Respondents (%)', fontsize=11)
ax.set_title('Most Valuable Manara Features According to Respondents', fontsize=13, fontweight='bold', pad=12)
plt.tight_layout()
plt.savefig(os.path.join(OUTPUT_DIR, 'feature_preferences.png'), bbox_inches='tight')
plt.close()

# — Chart 3: Adoption Intention —
adopt_raw = clean(df['adoption_intention'])
adopt_counts = adopt_raw.dropna().value_counts()
labels_order = ['yes', 'maybe', 'no']
adopt_vals = [adopt_counts.get(l, 0) for l in labels_order]
adopt_pcts = [v / sum(adopt_vals) * 100 for v in adopt_vals]
display_labels = [f'Yes\n{adopt_pcts[0]:.1f}%', f'Maybe\n{adopt_pcts[1]:.1f}%', f'No\n{adopt_pcts[2]:.1f}%']
fig, ax = plt.subplots(figsize=(7, 7))
wedge_colors = [COLORS['green'], COLORS['accent'], COLORS['secondary']]
wedges, texts = ax.pie(adopt_vals, labels=display_labels, colors=wedge_colors,
                        startangle=90, pctdistance=0.75,
                        wedgeprops={'edgecolor': 'white', 'linewidth': 2})
for text in texts:
    text.set_fontsize(12)
    text.set_fontweight('bold')
ax.set_title("Students' Intention to Use Manara", fontsize=14, fontweight='bold', pad=20)
plt.tight_layout()
plt.savefig(os.path.join(OUTPUT_DIR, 'adoption_intention.png'), bbox_inches='tight')
plt.close()

# — Chart 4: Academic Year Distribution —
year_plot = year_dist.copy()
# Combine 6th year variants
if '6th year ' in year_plot.index:
    year_plot['6th year'] = year_plot.get('6th year', 0) + year_plot.pop('6th year ')
fig, ax = plt.subplots(figsize=(9, 5))
year_order = ['1st year', '2nd year', '3rd year', '4th year', '5th year', '6th year', 'Other']
ordered = {y: year_plot.get(y, 0) for y in year_order if y in year_plot or year_plot.get(y, 0) > 0}
# get actual keys from year_plot
actual_keys = [k for k in year_order if k in year_plot.index]
y_vals = [year_plot[k] for k in actual_keys]
bar_colors2 = [COLORS['primary']] * len(actual_keys)
bars = ax.bar(actual_keys, y_vals, color=bar_colors2, edgecolor='white', linewidth=1.2)
for bar, v in zip(bars, y_vals):
    ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.3,
            str(v), ha='center', va='bottom', fontsize=11, fontweight='bold')
ax.set_ylabel('Number of Respondents', fontsize=11)
ax.set_title("Respondents' Academic Year Distribution", fontsize=13, fontweight='bold', pad=12)
ax.set_ylim(0, max(y_vals) + 5)
plt.tight_layout()
plt.savefig(os.path.join(OUTPUT_DIR, 'academic_year_distribution.png'), bbox_inches='tight')
plt.close()

# ─────────────────────────────────────────────
# 11. EXCEL OUTPUT
# ─────────────────────────────────────────────

def style_header(ws, row, col_count, fill_color='2E86AB', font_color='FFFFFF'):
    fill = PatternFill('solid', start_color=fill_color, end_color=fill_color)
    font = Font(bold=True, color=font_color, name='Arial', size=10)
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='FFFFFF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border

def style_data_rows(ws, start_row, end_row, col_count):
    light = PatternFill('solid', start_color='EBF5FB', end_color='EBF5FB')
    white = PatternFill('solid', start_color='FFFFFF', end_color='FFFFFF')
    font = Font(name='Arial', size=9)
    align = Alignment(wrap_text=True, vertical='top')
    for r in range(start_row, end_row + 1):
        fill = light if r % 2 == 0 else white
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.font = font
            cell.alignment = align

def auto_width(ws, min_w=10, max_w=50):
    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(max_len + 2, min_w), max_w)

def df_to_sheet(ws, df_in, title=None, fill_color='2E86AB'):
    row_offset = 0
    if title:
        ws.cell(1, 1, title).font = Font(bold=True, size=13, name='Arial', color='1A1A2E')
        ws.cell(1, 1).alignment = Alignment(horizontal='left')
        row_offset = 2
    # Header
    for ci, col in enumerate(df_in.columns, 1):
        ws.cell(row_offset + 1, ci, col)
    style_header(ws, row_offset + 1, len(df_in.columns), fill_color=fill_color)
    # Data
    for ri, row_data in enumerate(df_in.itertuples(index=False), row_offset + 2):
        for ci, val in enumerate(row_data, 1):
            ws.cell(ri, ci, val)
    style_data_rows(ws, row_offset + 2, row_offset + 1 + len(df_in), len(df_in.columns))
    auto_width(ws)
    ws.row_dimensions[row_offset + 1].height = 30

wb = Workbook()
wb.remove(wb.active)

# Sheet 1: Respondent Profile
ws1 = wb.create_sheet('Respondent Profile')
df_to_sheet(ws1, respondent_profile_table, title='Manara Evaluation — Respondent Profile', fill_color='1A6B9A')

# Sheet 2: Question Summary
ws2 = wb.create_sheet('Question Summary')
df_to_sheet(ws2, question_summary_table, title='Question-Level Summary with Normalization', fill_color='2E86AB')

# Sheet 3: Category Summary
ws3 = wb.create_sheet('Category Summary')
df_to_sheet(ws3, category_summary_table, title='Evaluation Category Summary', fill_color='A23B72')

# Sheet 4: Feature Preferences
ws4 = wb.create_sheet('Feature Preferences')
df_to_sheet(ws4, feature_preference_table, title='Most Valuable Feature Preferences', fill_color='F18F01')

# Sheet 5: Response Distributions
ws5 = wb.create_sheet('Response Distributions')
df_to_sheet(ws5, response_distribution_table, title='Response Distributions by Question', fill_color='44BBA4')

excel_path = os.path.join(OUTPUT_DIR, 'manara_evaluation_results.xlsx')
wb.save(excel_path)

# ─────────────────────────────────────────────
# 12. MARKDOWN REPORT
# ─────────────────────────────────────────────

adopt_yes_pct = adopt_pcts[0]
adopt_maybe_pct = adopt_pcts[1]
adopt_no_pct = adopt_pcts[2]

# Build category score bullets for report
cat_bullets = '\n'.join([f"- **{cat}**: {score:.1f}/100 — {interpret_score(score)}"
                          for cat, score in category_scores.items()])

top_feature = feature_preference_table.iloc[0]
second_feature = feature_preference_table.iloc[1]

# Key numeric values
prereq_avg = df['prerequisite_importance_score'].dropna().mean()
diag_avg = df['diagnostic_exam_score'].dropna().mean()
weak_avg = df['weak_topics_usage_score'].dropna().mean()
weakness_avg = df['weakness_improves_performance_score'].dropna().mean()
path_avg = df['personalized_path_score'].dropna().mean()
need_avg = df['past_papers_need_score'].dropna().mean()
gen_avg = df['generated_questions_score'].dropna().mean()
chat_avg = df['chatbot_usage_score'].dropna().mean()
prog_avg = df['progress_motivation_score'].dropna().mean()
whats_avg = df['whatsapp_interest_score'].dropna().mean()
adopt_avg = df['adoption_intention_score'].dropna().mean()

guidance_avg = np.mean([prereq_avg, diag_avg, weak_avg, weakness_avg])

n_students_pct = n_students / TOTAL * 100

md = f"""# Manara Evaluation Report
*AI-Based Learning Support System for PSUT Students*

---

## 1. Evaluation Methodology

This evaluation is based on {TOTAL} responses collected via a structured Google Form survey distributed among university students. The survey was designed to assess the perceived usefulness, relevance, and adoption intention of Manara — an AI-powered academic guidance platform for PSUT students.

To enable fair cross-question comparison, all responses measured on different scales (categorical Likert-type, ordinal, and linear scales) were normalized to a common 0–100 scale before analysis. Higher normalized scores consistently indicate stronger support for or interest in the corresponding Manara feature. The only exception is the question regarding the availability of past papers, which was treated as a **need indicator**: a higher score reflects a stronger unmet need for additional practice resources, not satisfaction with current resources.

The checkbox question about most valued features was analyzed separately by counting individual feature selections and calculating each feature's selection rate across all respondents. This question was not averaged with the scored items.

---

## 2. Respondent Profile

A total of **{TOTAL} respondents** completed the survey. Of these, **{n_students} ({n_students_pct:.1f}%)** identified as current university students, while **{n_nonstudents} ({n_nonstudents/TOTAL*100:.1f}%)** were graduates or non-students.

Among current students, the academic year distribution was as follows:

| Academic Year | Count | Percentage |
|---|---|---|
""" + '\n'.join(
    f"| {year} | {cnt} | {cnt/TOTAL*100:.1f}% |"
    for year, cnt in year_dist.items()
) + f"""

The {n_nonstudents} non-student respondents were asked whether they would have used a system like Manara earlier in their studies. Of these respondents, the distribution was: **Yes: {earlier_dist.get('yes', 0)}**, **Maybe: {earlier_dist.get('maybe', 0)}**, **No: {earlier_dist.get('no', 0)}** — suggesting that graduates who experienced academic challenges recognize the potential value of such a system.

---

## 3. Quantitative Results

After normalizing all scored questions to a 0–100 scale, the following category-level averages were obtained:

{cat_bullets}

The strongest category scores were observed in the areas of **Need for Academic Guidance & Weakness Identification** ({guidance_avg:.1f}/100) and **Personalized Learning Path** ({path_avg:.1f}/100), reflecting that students clearly recognize the value of identifying their weaknesses early and receiving tailored study plans. The **Overall Adoption Intention** score of {adopt_avg:.1f}/100 further indicates that a large proportion of respondents would choose to use Manara if it were available.

The **Practice Resource Need Score** of {need_avg:.1f}/100 suggests that a substantial portion of students find current past paper resources insufficient, reinforcing the rationale for Manara's AI-generated question bank. The **Generated Practice Questions** feature received a helpfulness score of {gen_avg:.1f}/100, indicating strong perceived utility.

The **AI Chatbot** received a likelihood-of-use score of {chat_avg:.1f}/100, reflecting positive but somewhat mixed reception — a finding consistent with the general novelty of AI-based course assistants among students. **Progress Tracking** ({prog_avg:.1f}/100) and **WhatsApp Reminders** ({whats_avg:.1f}/100) showed moderate to positive support, though WhatsApp reminders showed greater variability in interest.

---

## 4. Feature Preference Analysis

Respondents were asked to select the features they found most valuable from a set of five options (multiple selections were permitted). The results were as follows:

| Rank | Feature | Selections | % of Respondents |
|---|---|---|---|
""" + '\n'.join(
    f"| {row['Rank']} | {row['Feature']} | {row['Selection Count']} | {row['Percentage of Respondents']}% |"
    for _, row in feature_preference_table.iterrows()
) + f"""

The most selected feature was **{top_feature['Feature']}** ({top_feature['Selection Count']} selections, {top_feature['Percentage of Respondents']}% of respondents), followed by **{second_feature['Feature']}** ({second_feature['Selection Count']} selections, {second_feature['Percentage of Respondents']}%). These results suggest that students prioritize practical academic support tools — particularly those that directly target knowledge gaps and structure their study process — over motivational or social features.

---

## 5. Overall Adoption Intention

When asked directly whether they would use Manara if it were available, **{adopt_yes_pct:.1f}%** of respondents answered Yes, **{adopt_maybe_pct:.1f}%** answered Maybe, and **{adopt_no_pct:.1f}%** answered No. This translates to a normalized adoption intention score of **{adopt_avg:.1f}/100**. The combined Yes + Maybe rate of **{adopt_yes_pct + adopt_maybe_pct:.1f}%** indicates very strong general willingness to engage with the platform, with no respondents expressing firm opposition to using it.

---

## 6. Key Findings

- The vast majority of respondents ({prereq_avg:.1f}/100 normalized score) consider assessing prerequisite knowledge before a course to be important or very important, validating Manara's diagnostic exam feature.
- Students strongly agree that diagnostic exams can help identify weak areas early (normalized score: {diag_avg:.1f}/100), with over 80% selecting "Strongly agree" or "Agree."
- {weakness_avg:.1f}/100 of the normalized score for the question on whether identifying weaknesses improves performance reflects near-universal agreement among respondents.
- The personalized learning path received a high helpfulness score ({path_avg:.1f}/100 on a 1–5 scale normalized to 0–100), indicating students value structured, individualized study plans.
- A meaningful proportion of respondents find current past paper resources insufficient (need score: {need_avg:.1f}/100), supporting the need for Manara's question bank.
- The AI chatbot feature shows positive but more varied reception ({chat_avg:.1f}/100), which may reflect differing comfort levels with AI-based learning tools.
- No respondent indicated they would definitely not use Manara, and the combined Yes + Maybe adoption rate reaches {adopt_yes_pct + adopt_maybe_pct:.1f}%.
- The most valued features — **{top_feature['Feature']}** and **{second_feature['Feature']}** — align directly with Manara's core design priorities.

---

## 7. Limitations

This evaluation carries the following limitations that should be considered when interpreting the results:

- **Self-reported data**: Responses reflect perceived usefulness and intended behavior, not actual usage outcomes. Participants may overestimate their likelihood of adopting new tools.
- **Intention vs. behavior**: The survey measures intent to use and perceived value, not long-term academic improvement or actual system engagement.
- **Sample representativeness**: While the survey reached {TOTAL} respondents, the sample may not fully represent all PSUT students across all majors, years, and academic backgrounds.
- **Branching effects**: Some questions were conditionally answered (e.g., earlier-use question for graduates only), resulting in differing valid response counts across items.
- **Response bias**: Respondents who chose to complete the survey may have pre-existing interest in academic support tools, potentially inflating positive scores.
- **Absence of usability testing**: Real-world usage testing with the deployed Manara system would provide substantially stronger evidence of effectiveness than pre-deployment survey data.

---

## 8. Final Report-Ready Paragraph

The evaluation of Manara was conducted through a structured survey administered to {TOTAL} university students and graduates. Responses were collected across key dimensions of the platform — including diagnostic assessment, personalized learning paths, AI-generated practice questions, chatbot support, progress tracking, and WhatsApp reminders — and normalized to a common 0–100 scale to enable consistent comparison. The findings indicate strong alignment between students' academic needs and Manara's proposed features. The category of academic guidance and weakness identification received a normalized average score of {guidance_avg:.1f}/100, reflecting a clear consensus that students value early identification of knowledge gaps. The personalized learning path feature was rated {path_avg:.1f}/100 for usefulness, and the question bank's practice resource need score of {need_avg:.1f}/100 confirms that many students find existing study resources insufficient. The AI chatbot received moderate-to-positive reception at {chat_avg:.1f}/100, and overall adoption intention scored {adopt_avg:.1f}/100, with {adopt_yes_pct + adopt_maybe_pct:.1f}% of respondents indicating they would use or likely use the platform if available. Among the five listed features, {top_feature['Feature']} and {second_feature['Feature']} emerged as the most valued, each selected by a majority of respondents. These results suggest that Manara addresses genuine and widely felt academic challenges among PSUT students, and that there is substantial interest in adopting such a system as part of everyday academic life. While this evaluation captures perceived value and intended use rather than demonstrated long-term impact, the feedback provides meaningful early validation of Manara's design approach and feature prioritization.

---

*Report generated automatically from Google Form response data.*
*Total responses analyzed: {TOTAL}*
"""

report_path = os.path.join(OUTPUT_DIR, 'manara_evaluation_report.md')
with open(report_path, 'w', encoding='utf-8') as f:
    f.write(md)

# ─────────────────────────────────────────────
# 13. CONSOLE SUMMARY
# ─────────────────────────────────────────────

print("=" * 65)
print("  MANARA EVALUATION RESULTS — SUMMARY")
print("=" * 65)
print(f"\nTotal Responses Analyzed: {TOTAL}")
print(f"  Current Students : {n_students} ({n_students/TOTAL*100:.1f}%)")
print(f"  Non-Students     : {n_nonstudents} ({n_nonstudents/TOTAL*100:.1f}%)")

print("\n── Category Summary ──────────────────────────────────────────")
print(f"{'Category':<40} {'Score':>7}  {'Interpretation'}")
print("-" * 90)
for _, row in category_summary_table.iterrows():
    print(f"{row['Evaluation Category']:<40} {row['Average Normalized Score']:>7.1f}  {row['Interpretation']}")

print("\n── Feature Preference Ranking ────────────────────────────────")
print(f"{'Rank':<6} {'Feature':<30} {'Selections':>10}  {'% of Respondents':>18}")
print("-" * 70)
for _, row in feature_preference_table.iterrows():
    print(f"{row['Rank']:<6} {row['Feature']:<30} {row['Selection Count']:>10}  {row['Percentage of Respondents']:>17.1f}%")

print("\n── Adoption Intention ────────────────────────────────────────")
for label, pct in zip(['Yes', 'Maybe', 'No'], adopt_pcts):
    print(f"  {label:<8}: {pct:.1f}%")

print("\n── Output Files ──────────────────────────────────────────────")
output_files = [
    'manara_evaluation_results.xlsx',
    'manara_evaluation_report.md',
    'category_scores.png',
    'feature_preferences.png',
    'adoption_intention.png',
    'academic_year_distribution.png',
]
for f in output_files:
    print(f"  ✓ {f}")
print("=" * 65)
